"""
exporter.py
Excel (.xlsx) and plaintext (.txt) export logic for the TLOPO Loot Tracker.
"""

import os
import sqlite3
from datetime import datetime
from typing import Optional

from enrichment import CLIENT_DB_FIELDS, CLIENT_DB_PROVENANCE_FIELDS, compute_session_statistics, enrich_events
from loot_parser import RARITY_ORDER
from session import Session

RARITY_FILL_HEX = {
    "Crude": "D9D9D9",
    "Common": "FFF2A8",
    "Rare": "C6EFCE",
    "Famed": "BDD7EE",
    "Legendary": "FFC7CE",
}
RARITY_FONT_HEX = {
    "Crude": "595959",
    "Common": "9C6500",
    "Rare": "006100",
    "Famed": "1F4E78",
    "Legendary": "9C0006",
}

RARITY_SORT_INDEX = {"Legendary": 0, "Famed": 1, "Rare": 2, "Common": 3, "Crude": 4}


def default_export_folder() -> str:
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    folder = os.path.join(desktop, "TLOPO_Tracker_Exports")
    return folder


def _timestamped_filename(prefix: str, ext: str) -> str:
    now = datetime.now()
    return f"{prefix}_{now.strftime('%Y-%m-%d_%H-%M')}.{ext}"


def export_to_excel(session: Session, folder: Optional[str] = None) -> str:
    """Write the 3-sheet Excel workbook. Returns the full path written."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is not installed. Run install.bat to set up dependencies."
        ) from e

    folder = folder or default_export_folder()
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, _timestamped_filename("TLOPO_Session", "xlsx"))

    wb = Workbook()

    # -----------------------------------------------------------------
    # Sheet 1 - Session Summary
    # -----------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Session Summary"
    headers1 = ["Target", "Kills", "Pouches", "Chests", "Skull Chests", "Skull Rate",
                "Crude", "Common", "Rare", "Famed", "Legendary"]
    ws1.append(headers1)
    for c in ws1[1]:
        c.font = Font(bold=True)

    total = session.session_totals()
    for name, stats in sorted(session.targets.items()):
        row = [
            stats.name, stats.kills, stats.pouches, stats.chests, stats.skull_chests,
            f"{stats.skull_rate():.1f}%",
            stats.rarity_counts.get("Crude", 0),
            stats.rarity_counts.get("Common", 0),
            stats.rarity_counts.get("Rare", 0),
            stats.rarity_counts.get("Famed", 0),
            stats.rarity_counts.get("Legendary", 0),
        ]
        ws1.append(row)
        r = ws1.max_row
        if stats.rarity_counts.get("Legendary", 0) > 0:
            for cell in ws1[r]:
                cell.fill = PatternFill("solid", fgColor=RARITY_FILL_HEX["Legendary"])
        elif stats.rarity_counts.get("Famed", 0) > 0:
            for cell in ws1[r]:
                cell.fill = PatternFill("solid", fgColor=RARITY_FILL_HEX["Famed"])

    ws1.append([
        "TOTAL", total.kills, total.pouches, total.chests, total.skull_chests,
        f"{total.skull_rate():.1f}%",
        total.rarity_counts.get("Crude", 0),
        total.rarity_counts.get("Common", 0),
        total.rarity_counts.get("Rare", 0),
        total.rarity_counts.get("Famed", 0),
        total.rarity_counts.get("Legendary", 0),
    ])
    for cell in ws1[ws1.max_row]:
        cell.font = Font(bold=True)

    for i, h in enumerate(headers1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 4)

    # -----------------------------------------------------------------
    # Sheet 2 - Named Item Log
    # -----------------------------------------------------------------
    ws2 = wb.create_sheet("Named Item Log")
    headers2 = ["Item Name", "Rarity", "Times Obtained", "Targets Found From"]
    ws2.append(headers2)
    for c in ws2[1]:
        c.font = Font(bold=True)

    named_records = sorted(
        session.named_items.values(),
        key=lambda r: (RARITY_SORT_INDEX.get(r.rarity, 99), -r.count),
    )
    for rec in named_records:
        ws2.append([rec.name, rec.rarity, rec.count, ", ".join(sorted(rec.targets))])
        r = ws2.max_row
        fill_hex = RARITY_FILL_HEX.get(rec.rarity)
        if fill_hex:
            for cell in ws2[r]:
                cell.fill = PatternFill("solid", fgColor=fill_hex)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 40

    # -----------------------------------------------------------------
    # Sheet 3 - Raw Events (one row per kill or per loot item, in the
    # order they happened -- the immutable per-observation record; see
    # session.py Session.raw_events). Placed before "Full Loot Log" so
    # it's the first thing a researcher used to the old sheet order
    # notices; "Full Loot Log" and the sheets after it are kept as-is
    # for backward compatibility with anything already parsing them.
    # -----------------------------------------------------------------
    ws_raw = wb.create_sheet("Raw Events")
    client_db_headers = ["Boss Group", "Enemy Base Type", "Expected Rare Chest %",
                          "Enemy Family", "Known Level Range", "Known Boss Status"]
    # Evidence-status columns ride alongside the client-DB columns rather
    # than being folded into them, so a reader can see at a glance
    # whether the row's Boss Group/Base Type/etc. values are "verified"
    # or merely "reported"/"inferred" -- see enrichment.CLIENT_DB_LOOKUP.
    evidence_headers = ["Evidence Status", "Evidence Source", "Evidence Notes"]
    headers_raw = ["Observation ID", "Timestamp", "Event Type", "Target", "Enemy Color",
                    "Location", "Associated Kill #", "Linked Kill Observation ID", "Link Status",
                    "Capture Quality", "Chest Type", "Item Name", "Item Category", "Category Group",
                    "Rarity", "OCR Confidence", "Confidence Tier", "Gold", "Chest Correlation ID"] \
        + client_db_headers + evidence_headers
    ws_raw.append(headers_raw)
    for c in ws_raw[1]:
        c.font = Font(bold=True)
    RARITY_COL = 15  # 1-indexed "Rarity" column, for per-row fill/font below

    for event in enrich_events(session):
        base = [
            event.get("observation_id", ""), event.get("timestamp", ""),
            event.get("event_type", ""), event.get("target", ""),
            event.get("enemy_color") or "", event.get("location") or "",
            event.get("associated_kill_number", event.get("kill_number", 0)),
            event.get("linked_kill_observation_id") or "",
            event.get("link_status") or "",
        ]
        client_db_values = [event.get(f) for f in
                             ["boss_group", "enemy_base_type", "expected_rare_chest_pct",
                              "enemy_family", "known_level_range", "known_boss_status"]]
        client_db_values = [v if v is not None else "" for v in client_db_values]
        evidence_values = [event.get(f) or "" for f in CLIENT_DB_PROVENANCE_FIELDS]
        client_db_values = client_db_values + evidence_values

        if event.get("event_type") == "kill":
            ws_raw.append(base + [event.get("capture_quality", ""), "", "", "", "", "",
                                   "", "", "", ""] + client_db_values)
            continue

        items = event.get("items", [])
        gold = event.get("gold", 0)
        chest_correlation_id = event.get("session_id", "")
        chest_type = event.get("chest_type", "")
        capture_quality = event.get("capture_quality", "")
        if not items:
            ws_raw.append(base + [capture_quality, chest_type, "(no items read)", "", "", "",
                                   "", "", gold, chest_correlation_id] + client_db_values)
            continue
        for item in items:
            conf = item.get("name_confidence")
            conf_str = f"{conf:.0f}%" if conf is not None else ""
            ws_raw.append(base + [
                capture_quality, chest_type, item.get("name", ""), item.get("category") or "",
                item.get("category_group") or "", item.get("rarity") or "", conf_str,
                item.get("confidence_tier") or "", gold, chest_correlation_id,
            ] + client_db_values)
            r = ws_raw.max_row
            rarity = item.get("rarity") or ""
            fill_hex = RARITY_FILL_HEX.get(rarity)
            font_hex = RARITY_FONT_HEX.get(rarity)
            if fill_hex:
                ws_raw.cell(row=r, column=RARITY_COL).fill = PatternFill("solid", fgColor=fill_hex)
            if font_hex:
                ws_raw.cell(row=r, column=RARITY_COL).font = Font(color=font_hex, bold=(rarity in ("Famed", "Legendary")))

    for i, h in enumerate(headers_raw, start=1):
        ws_raw.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 4)

    # -----------------------------------------------------------------
    # Sheet 4 - Item Inventory (every item seen this session, aggregated
    # by name -- see Session.session_item_counts). Answers "how many
    # Rubies have we observed" without parsing free text.
    # -----------------------------------------------------------------
    ws_inv = wb.create_sheet("Item Inventory")
    headers_inv = ["Item Name", "Rarity", "Category", "Category Group", "Count"]
    ws_inv.append(headers_inv)
    for c in ws_inv[1]:
        c.font = Font(bold=True)

    inventory = sorted(
        session.session_item_counts().values(),
        key=lambda v: (RARITY_SORT_INDEX.get(v["rarity"], 99), -v["count"]),
    )
    for entry in inventory:
        ws_inv.append([entry["name"], entry["rarity"] or "", entry["category"] or "",
                        entry.get("category_group") or "", entry["count"]])
        r = ws_inv.max_row
        fill_hex = RARITY_FILL_HEX.get(entry["rarity"])
        if fill_hex:
            for cell in ws_inv[r]:
                cell.fill = PatternFill("solid", fgColor=fill_hex)

    for i, h in enumerate(headers_inv, start=1):
        ws_inv.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    # -----------------------------------------------------------------
    # Sheet 5 - Statistics (observed rates + Wilson 95% CIs -- see
    # enrichment.compute_session_statistics). One row per metric so this
    # stays easy to chart/pivot directly in Excel.
    # -----------------------------------------------------------------
    ws_stats = wb.create_sheet("Statistics")
    headers_stats = ["Metric", "Count", "Total", "Rate %", "95% CI Low %", "95% CI High %"]
    ws_stats.append(headers_stats)
    for c in ws_stats[1]:
        c.font = Font(bold=True)

    stats_summary = compute_session_statistics(session)
    kills = stats_summary["kills"]

    def _append_rate_row(metric: str, count: int, total: int, rate_ci: dict):
        rate = rate_ci.get("rate")
        lo, hi = rate_ci.get("ci_low"), rate_ci.get("ci_high")
        ws_stats.append([
            metric, count, total,
            f"{rate * 100:.1f}" if rate is not None else "",
            f"{lo * 100:.1f}" if lo is not None else "",
            f"{hi * 100:.1f}" if hi is not None else "",
        ])

    ws_stats.append(["Kills", kills, "", "", "", ""])
    _append_rate_row("Container Rate", stats_summary["containers"], kills, stats_summary["container_rate"])
    for label, dist in stats_summary["container_distribution"].items():
        pct = dist["pct_of_containers"]
        ws_stats.append([f"Container Distribution: {label.title()}", dist["count"],
                          stats_summary["containers"], f"{pct:.1f}" if pct is not None else "", "", ""])
    _append_rate_row("Skull Chest Rate", total.skull_chests, kills, stats_summary["skull_chest_rate"])
    for rarity in RARITY_ORDER:
        _append_rate_row(f"{rarity} Item Rate", total.rarity_counts.get(rarity, 0), kills,
                          stats_summary["rarity_rates"][rarity])

    for i, h in enumerate(headers_stats, start=1):
        ws_stats.column_dimensions[get_column_letter(i)].width = max(16, len(h) + 4)

    # -----------------------------------------------------------------
    # Sheet 6 - Full Loot Log
    # -----------------------------------------------------------------
    ws3 = wb.create_sheet("Full Loot Log")
    headers3 = ["Timestamp", "Target", "Chest Type", "Item Name", "Rarity", "Gold", "Kill Number"]
    ws3.append(headers3)
    for c in ws3[1]:
        c.font = Font(bold=True)

    for target_name, stats in sorted(session.targets.items()):
        for entry in stats.loot_log:
            items = entry.get("items", [])
            if not items:
                ws3.append([
                    entry.get("timestamp", ""), entry.get("target", target_name),
                    entry.get("chest_type", ""), "(no items)", "", entry.get("gold", 0),
                    entry.get("kill_number", 0),
                ])
                continue
            for item in items:
                ws3.append([
                    entry.get("timestamp", ""),
                    entry.get("target", target_name),
                    entry.get("chest_type", ""),
                    item.get("name", ""),
                    item.get("rarity", ""),
                    entry.get("gold", 0),
                    entry.get("kill_number", 0),
                ])
                r = ws3.max_row
                rarity = item.get("rarity", "")
                fill_hex = RARITY_FILL_HEX.get(rarity)
                font_hex = RARITY_FONT_HEX.get(rarity)
                if fill_hex:
                    ws3.cell(row=r, column=5).fill = PatternFill("solid", fgColor=fill_hex)
                if font_hex:
                    ws3.cell(row=r, column=5).font = Font(color=font_hex, bold=(rarity in ("Famed", "Legendary")))

    for i, h in enumerate(headers3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    wb.save(path)
    return path


def _format_duration(seconds: float) -> str:
    seconds = int(seconds)
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def export_to_text(session: Session, folder: Optional[str] = None) -> str:
    """Write the plaintext session export. Returns the full path written."""
    folder = folder or default_export_folder()
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, _timestamped_filename("TLOPO_Session", "txt"))

    lines = []
    lines.append("=== TLOPO LOOT TRACKER SESSION ===")
    lines.append(f"Session ID: {session.session_id}")
    lines.append(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"Duration: {_format_duration(session.duration_seconds())}")
    lines.append("")

    for name, stats in sorted(session.targets.items()):
        lines.append(f"--- {name.upper()} ({stats.kills} kills) ---")
        lines.append(
            f"Pouches: {stats.pouches} | Chests: {stats.chests} | "
            f"Skull Chests: {stats.skull_chests} ({stats.skull_rate():.1f}% skull rate)"
        )
        lines.append(
            "Crude: {c} | Common: {u} | Rare: {r} | Famed: {f} | Legendary: {l}".format(
                c=stats.rarity_counts.get("Crude", 0),
                u=stats.rarity_counts.get("Common", 0),
                r=stats.rarity_counts.get("Rare", 0),
                f=stats.rarity_counts.get("Famed", 0),
                l=stats.rarity_counts.get("Legendary", 0),
            )
        )
        lines.append("")

        target_famed = [r for r in session.named_items.values()
                         if r.rarity == "Famed" and name in r.targets]
        target_legendary = [r for r in session.named_items.values()
                             if r.rarity == "Legendary" and name in r.targets]
        target_famed.sort(key=lambda r: r.count, reverse=True)
        target_legendary.sort(key=lambda r: r.count, reverse=True)

        famed_total = sum(r.count for r in target_famed)
        lines.append(f"FAMED DROPS ({famed_total} total):")
        if target_famed:
            for rec in target_famed:
                lines.append(f"  {rec.name} x{rec.count}")
        else:
            lines.append("  None")
        lines.append("")

        legendary_total = sum(r.count for r in target_legendary)
        lines.append(f"LEGENDARY DROPS ({legendary_total} total):")
        if target_legendary:
            for rec in target_legendary:
                lines.append(f"  {rec.name} x{rec.count}")
        else:
            lines.append("  None")
        lines.append("")

        if stats.loot_log:
            lines.append("Full Loot Log:")
            for entry in stats.loot_log:
                items = entry.get("items", [])
                item_strs = [
                    f"{it['name']} ({it['rarity']})" if it.get("rarity") else it['name']
                    for it in items
                ] or ["(no items read)"]
                lines.append(
                    f"  [{entry.get('timestamp','')}] [{entry.get('target', name)}] "
                    f"[{entry.get('chest_type','')}] — {' '.join(item_strs)} — {entry.get('gold',0)}g"
                )
            lines.append("")

    total = session.session_totals()
    lines.append("=== SESSION TOTALS ===")
    lines.append(
        f"Total Kills: {total.kills} | Total Skull Chests: {total.skull_chests} "
        f"({total.skull_rate():.1f}% rate)"
    )
    lines.append(
        "Crude: {c} | Common: {u} | Rare: {r} | Famed: {f} | Legendary: {l}".format(
            c=total.rarity_counts.get("Crude", 0),
            u=total.rarity_counts.get("Common", 0),
            r=total.rarity_counts.get("Rare", 0),
            f=total.rarity_counts.get("Famed", 0),
            l=total.rarity_counts.get("Legendary", 0),
        )
    )
    lines.append("")

    all_famed = session.named_items_by_rarity("Famed")
    all_famed_total = sum(r.count for r in all_famed)
    lines.append(f"ALL FAMED DROPS THIS SESSION ({all_famed_total} total):")
    if all_famed:
        for rec in all_famed:
            lines.append(f"  {rec.name} x{rec.count}")
    else:
        lines.append("  None")
    lines.append("")

    all_legendary = session.named_items_by_rarity("Legendary")
    all_legendary_total = sum(r.count for r in all_legendary)
    lines.append(f"ALL LEGENDARY DROPS THIS SESSION ({all_legendary_total} total):")
    if all_legendary:
        for rec in all_legendary:
            lines.append(f"  {rec.name} x{rec.count}")
    else:
        lines.append("  None")
    lines.append("")

    # Observed rates + Wilson 95% CIs (see enrichment.compute_session_statistics)
    # -- separates P(any container | kill), P(container type | container),
    # and P(rarity tier | kill) into distinct numbers rather than one
    # collapsed "drop rate", per the collaborator's own container/loot
    # probability model.
    def _fmt_rate_ci(rate_ci: dict) -> str:
        rate, lo, hi = rate_ci.get("rate"), rate_ci.get("ci_low"), rate_ci.get("ci_high")
        if rate is None:
            return "n/a (no kills logged yet)"
        return f"{rate * 100:.1f}%  (95% Wilson CI: {lo * 100:.1f}%-{hi * 100:.1f}%)"

    stats_summary = compute_session_statistics(session)
    lines.append("=== OBSERVED STATISTICS ===")
    lines.append(f"Kills: {stats_summary['kills']}")
    lines.append(f"Containers: {stats_summary['containers']}")
    lines.append(f"Container Rate: {_fmt_rate_ci(stats_summary['container_rate'])}")
    lines.append("Container Distribution:")
    for label, dist in stats_summary["container_distribution"].items():
        pct = dist["pct_of_containers"]
        pct_str = f"{pct:.1f}%" if pct is not None else "n/a"
        lines.append(f"  {label.title()}: {dist['count']} ({pct_str} of containers)")
    lines.append(f"Skull Chest Rate: {_fmt_rate_ci(stats_summary['skull_chest_rate'])}")
    lines.append("Item Rarity Rates (per kill):")
    for rarity in RARITY_ORDER:
        lines.append(f"  {rarity}: {_fmt_rate_ci(stats_summary['rarity_rates'][rarity])}")
    lines.append("Expected Client Model: (not yet available -- see enrichment.CLIENT_DB_LOOKUP)")
    lines.append("")

    # Full item inventory (every item this session, not just Famed/
    # Legendary -- see Session.session_item_counts).
    inventory = sorted(
        session.session_item_counts().values(),
        key=lambda v: (RARITY_SORT_INDEX.get(v["rarity"], 99), -v["count"]),
    )
    lines.append(f"=== ITEM INVENTORY ({len(inventory)} unique item(s)) ===")
    if inventory:
        for entry in inventory:
            rarity_str = f" [{entry['rarity']}]" if entry["rarity"] else ""
            category_str = f" ({entry['category']})" if entry["category"] else ""
            lines.append(f"  {entry['name']}{rarity_str}{category_str} x{entry['count']}")
    else:
        lines.append("  None")
    lines.append("")

    # One block per raw observation (kill or loot event), in the order
    # they happened -- see session.py Session.raw_events for the
    # immutable, write-time record, and enrichment.py for the derived
    # kill<->loot linkage/capture-quality/client-DB fields computed here
    # at export time. Every summary section above this is a rollup
    # derived from this stream.
    lines.append("=== RAW EVENTS ===")
    enriched = enrich_events(session)
    if enriched:
        for event in enriched:
            is_kill = event.get("event_type") == "kill"
            lines.append("Kill Event" if is_kill else "Loot Event")
            lines.append(f"Observation ID: {event.get('observation_id', '')}")
            if is_kill:
                lines.append(f"Kill #: {event.get('kill_number', 0)}")
            else:
                lines.append(f"Associated Kill #: {event.get('associated_kill_number', 0)}")
                lines.append(f"Link Status: {event.get('link_status', 'unlinked')}")
                if event.get("linked_kill_observation_id"):
                    lines.append(f"Linked Kill Observation: {event['linked_kill_observation_id']}")
            lines.append(f"Enemy: {event.get('target', '')}")
            if event.get("enemy_color"):
                lines.append(f"Color: {event['enemy_color']}")
            if event.get("location"):
                lines.append(f"Location: {event['location']}")
            # Client-DB-sourced fields (see enrichment.CLIENT_DB_LOOKUP) --
            # only printed when non-blank, so today's output (an empty
            # lookup table) looks the same as before this join point
            # existed. Each is suffixed with its evidence_status (see
            # enrichment.CLIENT_DB_PROVENANCE_FIELDS) so a "reported" or
            # "inferred" value is never presented indistinguishably from
            # a "verified" one -- required by this project's evidence-tier
            # rule (see the loot-tracker experimental-branch spec / RE
            # bible: "no export may silently convert Tier C or E into fact").
            evidence_status = event.get("evidence_status")
            for field_name, label in [
                ("boss_group", "Boss Group"), ("enemy_base_type", "Enemy Base Type"),
                ("expected_rare_chest_pct", "Expected Rare Chest %"),
                ("enemy_family", "Enemy Family"), ("known_level_range", "Known Level Range"),
                ("known_boss_status", "Known Boss Status"),
            ]:
                if event.get(field_name) is not None:
                    suffix = f" ({evidence_status})" if evidence_status else " (evidence status unknown)"
                    lines.append(f"{label}: {event[field_name]}{suffix}")
            if evidence_status is not None and event.get("evidence_source"):
                lines.append(f"Evidence Source: {event['evidence_source']}")
            if evidence_status is not None and event.get("evidence_notes"):
                lines.append(f"Evidence Notes: {event['evidence_notes']}")
            if is_kill:
                lines.append(f"Capture Quality: {event.get('capture_quality', '')}")
            else:
                lines.append(f"Container: {event.get('chest_type', '')}")
                lines.append(f"Capture Quality: {event.get('capture_quality', '')}")
                items = event.get("items", [])
                if items:
                    # One sub-block per item, with OCR confidence and its
                    # tier as their own labeled lines (kept separate, not
                    # fused into one "42%, Flag" string, so a later parser
                    # can filter/aggregate on either independently).
                    for it in items:
                        lines.append(f"Item: {it.get('name', '')}")
                        if it.get("rarity"):
                            lines.append(f"  Rarity: {it['rarity']}")
                        if it.get("category"):
                            group = it.get("category_group")
                            group_str = f" ({group})" if group else ""
                            lines.append(f"  Category: {it['category']}{group_str}")
                        conf = it.get("name_confidence")
                        if conf is not None:
                            lines.append(f"  OCR Confidence: {conf:.0f}%")
                        if it.get("confidence_tier"):
                            lines.append(f"  Confidence Class: {it['confidence_tier']}")
                else:
                    lines.append("Items: (no items read)")
                lines.append(f"Gold: {event.get('gold', 0)}g")
            lines.append(f"Timestamp: {event.get('timestamp', '')}")
            lines.append("")
    else:
        lines.append("  None")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return path


def export_to_sqlite(session: Session, folder: Optional[str] = None) -> str:
    """
    Write a normalized SQLite database of this session's observations.
    Uses only the standard library `sqlite3` module -- no new dependency
    for one export format, same precedent as kraken_ledger_client.py's own
    stdlib-only network calls.

    Schema v3 (see enrichment.py for where the derived columns come
    from) -- a thin `observations` envelope (one row per kill OR loot
    event) with `kill_events`/`loot_events` detail tables underneath it,
    plus `enemies`/`locations` reference tables and a `session_statistics`
    results table (see enrichment.compute_session_statistics):
      sessions(session_id PK, session_start, exported_at)
      enemies(enemy_id PK, display_name, boss_group, enemy_base_type,
        expected_rare_chest_pct, enemy_family, known_level_range,
        known_boss_status, evidence_status, evidence_source,
        evidence_notes)
      locations(location_id PK)
      observations(observation_id PK, session_id FK, event_type,
        timestamp, enemy_id FK, enemy_color, location_id FK)
      kill_events(observation_id PK/FK, kill_number, capture_quality)
      loot_events(observation_id PK/FK, associated_kill_number,
        linked_kill_observation_id FK, link_status, chest_type, gold,
        capture_quality, chest_correlation_id)
      items(id PK, observation_id FK -> loot_events, item_name, rarity,
        category, name_confidence, confidence_tier)
      session_statistics(session_id FK, metric, value, ci_low, ci_high)

    v3 adds enemies.evidence_status/evidence_source/evidence_notes (see
    enrichment.CLIENT_DB_PROVENANCE_FIELDS) and loot_events.link_status
    (see enrichment.enrich_events) -- an enemy row's boss_group/
    enemy_base_type/etc. columns must never be read as verified fact
    without checking evidence_status alongside them.

    A `Containers` table was deliberately NOT added -- a container has no
    existence independent of the one loot_events row that observed it, so
    a separate table would just be a 1:1 join adding no information; its
    chest_type/gold live directly on loot_events instead.

    session_statistics is written as a plain table (computed once here
    from enrichment.compute_session_statistics/stats.py), not a SQL VIEW
    -- SQLite's math-function support (needed for the Wilson interval's
    sqrt) isn't guaranteed present in every Python build's bundled
    SQLite, so the CI math stays in Python and only the results land in
    the database.

    This is the artifact meant for real analysis (SQL joins/aggregates,
    or a straight import into Postgres/pandas later) -- the Excel/text
    exports above stay the human-readable summaries.
    """
    folder = folder or default_export_folder()
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, _timestamped_filename("TLOPO_Session", "sqlite"))

    conn = sqlite3.connect(path)
    try:
        cur = conn.cursor()
        cur.execute("PRAGMA user_version = 3")  # schema generation marker -- v1 was the flat single-table observations/observation_items shape from the prior export round, v2 added the normalized tables below, v3 added evidence/link_status columns (see docstring above)

        cur.execute("CREATE TABLE sessions(session_id TEXT PRIMARY KEY, session_start REAL, exported_at REAL)")
        cur.execute("""
            CREATE TABLE enemies(
                enemy_id TEXT PRIMARY KEY,
                display_name TEXT,
                boss_group TEXT,
                enemy_base_type TEXT,
                expected_rare_chest_pct REAL,
                enemy_family TEXT,
                known_level_range TEXT,
                known_boss_status TEXT,
                evidence_status TEXT,
                evidence_source TEXT,
                evidence_notes TEXT
            )
        """)
        cur.execute("CREATE TABLE locations(location_id TEXT PRIMARY KEY)")
        cur.execute("""
            CREATE TABLE observations(
                observation_id TEXT PRIMARY KEY,
                session_id TEXT REFERENCES sessions(session_id),
                event_type TEXT,
                timestamp TEXT,
                enemy_id TEXT REFERENCES enemies(enemy_id),
                enemy_color TEXT,
                location_id TEXT REFERENCES locations(location_id)
            )
        """)
        cur.execute("""
            CREATE TABLE kill_events(
                observation_id TEXT PRIMARY KEY REFERENCES observations(observation_id),
                kill_number INTEGER,
                capture_quality TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE loot_events(
                observation_id TEXT PRIMARY KEY REFERENCES observations(observation_id),
                associated_kill_number INTEGER,
                linked_kill_observation_id TEXT REFERENCES observations(observation_id),
                link_status TEXT,
                chest_type TEXT,
                gold INTEGER,
                capture_quality TEXT,
                chest_correlation_id TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE items(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                observation_id TEXT REFERENCES loot_events(observation_id),
                item_name TEXT,
                rarity TEXT,
                category TEXT,
                category_group TEXT,
                name_confidence REAL,
                confidence_tier TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE session_statistics(
                session_id TEXT REFERENCES sessions(session_id),
                metric TEXT,
                value REAL,
                ci_low REAL,
                ci_high REAL
            )
        """)
        cur.execute("CREATE INDEX idx_obs_session ON observations(session_id)")
        cur.execute("CREATE INDEX idx_obs_enemy ON observations(enemy_id)")
        cur.execute("CREATE INDEX idx_items_obs ON items(observation_id)")

        cur.execute(
            "INSERT INTO sessions VALUES (?, ?, ?)",
            (session.session_id, session.session_start, datetime.now().timestamp()),
        )

        seen_enemy_ids = set()
        seen_location_ids = set()
        enriched = enrich_events(session)

        for event in enriched:
            observation_id = event.get("observation_id")
            if observation_id is None:
                # Pre-observation-ID data restored from an older
                # autosave -- skip rather than insert a NULL primary key.
                continue

            enemy_id = event.get("enemy_id") or None
            if enemy_id and enemy_id not in seen_enemy_ids:
                seen_enemy_ids.add(enemy_id)
                cur.execute(
                    "INSERT INTO enemies VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        enemy_id, event.get("target"),
                        *(event.get(f) for f in CLIENT_DB_FIELDS),
                        *(event.get(f) for f in CLIENT_DB_PROVENANCE_FIELDS),
                    ),
                )

            location_id = event.get("location") or None
            if location_id and location_id not in seen_location_ids:
                seen_location_ids.add(location_id)
                cur.execute("INSERT INTO locations VALUES (?)", (location_id,))

            cur.execute(
                "INSERT INTO observations VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    observation_id, session.session_id, event.get("event_type"),
                    event.get("timestamp"), enemy_id, event.get("enemy_color"), location_id,
                ),
            )

            if event.get("event_type") == "kill":
                cur.execute(
                    "INSERT INTO kill_events VALUES (?, ?, ?)",
                    (observation_id, event.get("kill_number"), event.get("capture_quality")),
                )
            else:
                cur.execute(
                    "INSERT INTO loot_events VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        observation_id, event.get("associated_kill_number"),
                        event.get("linked_kill_observation_id"), event.get("link_status"),
                        event.get("chest_type"),
                        event.get("gold"), event.get("capture_quality"), event.get("session_id"),
                    ),
                )
                for item in event.get("items", []):
                    cur.execute(
                        "INSERT INTO items (observation_id, item_name, rarity, category, category_group, "
                        "name_confidence, confidence_tier) VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (
                            observation_id, item.get("name"), item.get("rarity"),
                            item.get("category"), item.get("category_group"),
                            item.get("name_confidence"), item.get("confidence_tier"),
                        ),
                    )

        stats_summary = compute_session_statistics(session)

        def _stat_row(metric: str, rate_ci: dict):
            cur.execute(
                "INSERT INTO session_statistics VALUES (?, ?, ?, ?, ?)",
                (session.session_id, metric, rate_ci.get("rate"), rate_ci.get("ci_low"), rate_ci.get("ci_high")),
            )

        cur.execute(
            "INSERT INTO session_statistics VALUES (?, 'kills', ?, NULL, NULL)",
            (session.session_id, stats_summary["kills"]),
        )
        _stat_row("container_rate", stats_summary["container_rate"])
        _stat_row("skull_chest_rate", stats_summary["skull_chest_rate"])
        for label, dist in stats_summary["container_distribution"].items():
            pct = dist["pct_of_containers"]
            cur.execute(
                "INSERT INTO session_statistics VALUES (?, ?, ?, NULL, NULL)",
                (session.session_id, f"container_distribution_{label}", pct / 100.0 if pct is not None else None),
            )
        for rarity in RARITY_ORDER:
            _stat_row(f"rarity_rate_{rarity.lower()}", stats_summary["rarity_rates"][rarity])

        conn.commit()
    finally:
        conn.close()

    return path
